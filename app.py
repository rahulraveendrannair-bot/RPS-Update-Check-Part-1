import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime
import requests
from bs4 import BeautifulSoup

st.set_page_config(
    page_title="RPS-BOT",
    page_icon="🤖",
    layout="wide"
)

st.title("🤖 RPS-BOT Dashboard")
st.caption(f"Last refreshed: {datetime.now().strftime('%d %B %Y, %H:%M:%S')}")
st.markdown("---")

# ── Scraper Functions (No Selenium — requests + BeautifulSoup only) ─────────

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}

def scrape_GB_FCD_UK_SANCTIONS_LIST():
    url = "https://www.gov.uk/government/publications/the-uk-sanctions-list"
    r = requests.get(url, headers=HEADERS, timeout=15)
    soup = BeautifulSoup(r.text, "html.parser")
    result = soup.find_all("dd", class_="gem-c-metadata__definition")[2].text.strip()
    return result.split("—")[0].strip()

def scrape_HM_Treasury_Consolidated_List():
    url = "https://www.gov.uk/government/publications/financial-sanctions-consolidated-list-of-targets"
    r = requests.get(url, headers=HEADERS, timeout=15)
    soup = BeautifulSoup(r.text, "html.parser")
    result = soup.find_all("dd", class_="gem-c-metadata__definition")[2].text.strip()
    return result.split("—")[0].strip()

def scrape_US_DOJ_HUMANTRAFFICKING_CASES():
    url = "https://www.justice.gov/humantrafficking/press-room"
    r = requests.get(url, headers=HEADERS, timeout=15)
    soup = BeautifulSoup(r.text, "html.parser")
    return soup.find("time").text.strip()

def scrape_SG_MAS_IAL():
    url = "https://www.mas.gov.sg/investor-alert-list?page=1&rows=All"
    r = requests.get(url, headers=HEADERS, timeout=15)
    soup = BeautifulSoup(r.text, "html.parser")
    return soup.find("div", class_="mas-search-card__meta").text.split(":")[1].strip()

def scrape_HK_SFC_AL():
    url = "https://www.sfc.hk/en/alert-list"
    r = requests.get(url, headers=HEADERS, timeout=15)
    soup = BeautifulSoup(r.text, "html.parser")
    table = soup.find("div", class_="table-container main-style alert-list-table")
    table1 = table.find_all("tr")[1]
    return table1.find_all("td")[2].text.strip()

SCRAPER_MAP = {
    "GB_FCD_UK_SANCTIONS_LIST":      scrape_GB_FCD_UK_SANCTIONS_LIST,
    "HM_Treasury_Consolidated_List": scrape_HM_Treasury_Consolidated_List,
    "US_DOJ_HUMANTRAFFICKING_CASES": scrape_US_DOJ_HUMANTRAFFICKING_CASES,
    "SG_MAS_IAL":                    scrape_SG_MAS_IAL,
    "HK_SFC_AL":                     scrape_HK_SFC_AL,
}

def run_scraper(func_name):
    try:
        func = SCRAPER_MAP.get(func_name)
        if func is None:
            return None, f"No scraper found for {func_name}"
        result = func()
        return result, None
    except Exception as e:
        return None, str(e)

# ── Status Styling ──────────────────────────────────────────────────────────
def style_status(val):
    if val == "Updated":
        return "background-color: #d4edda; color: #155724; font-weight: bold;"
    elif val == "Not Updated":
        return "background-color: #d6d8db; color: #383d41; font-weight: bold;"
    elif val == "Failed":
        return "background-color: #f8d7da; color: #721c24; font-weight: bold;"
    return ""

# ── File Upload ─────────────────────────────────────────────────────────────
st.subheader("📤 Step 1 — Upload RPS-BOT.xlsx")
uploaded_file = st.file_uploader("Upload your RPS-BOT.xlsx file", type=["xlsx"])

if uploaded_file:
    file_bytes = uploaded_file.read()

    # Load workbook preserving hyperlinks
    wb = load_workbook(BytesIO(file_bytes))
    ws = wb.active

    # Extract hyperlinks from column B
    hyperlinks = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell = row[1]
        hyperlinks.append(cell.hyperlink.target if cell.hyperlink else None)

    # Load into DataFrame
    df = pd.read_excel(BytesIO(file_bytes), dtype=str)

    # Ensure all required columns exist
    for col in ["Previous Data", "Current Data", "Status", "Tracking Number", "Creation Date"]:
        if col not in df.columns:
            df[col] = None

    st.success(f"✅ File uploaded — {len(df)} sources found")

    # ── Show current data ───────────────────────────────────────────────────
    st.subheader("📊 Current Data")
    st.dataframe(
        df.style.applymap(style_status, subset=["Status"]),
        use_container_width=True,
        height=230
    )

    st.markdown("---")

    # ── Run Scrapers Button ─────────────────────────────────────────────────
    st.subheader("▶️ Step 2 — Run Scrapers")

    if st.button("🚀 Run Scrapers Now", type="primary", use_container_width=True):

        # Shift Current Data → Previous Data
        df["Previous Data"] = df["Current Data"]

        st.info("⏳ Scrapers running... please wait.")

        progress_bar = st.progress(0)
        status_log   = st.empty()
        results_area = st.empty()

        total = len(df)
        logs  = []

        for i in df.index:
            func_name = df.loc[i, "RPL-TYPE"]
            status_log.markdown(f"🔄 Running: **{func_name}**")

            output_date, error = run_scraper(func_name)

            if error:
                df.loc[i, "Status"]       = "Failed"
                df.loc[i, "Current Data"] = None
                logs.append(f"❌ **{func_name}** — Failed: {error}")
            else:
                df.loc[i, "Current Data"] = str(output_date)
                if df.loc[i, "Previous Data"] == df.loc[i, "Current Data"]:
                    df.loc[i, "Status"] = "Not Updated"
                    logs.append(f"🔵 **{func_name}** — Not Updated ({output_date})")
                else:
                    df.loc[i, "Status"] = "Updated"
                    logs.append(f"🟢 **{func_name}** — Updated! `{df.loc[i, 'Previous Data']}` → `{output_date}`")

            progress_bar.progress((i + 1) / total)
            results_area.markdown("\n\n".join(logs))

        status_log.markdown("✅ **All scrapers completed!**")

        # ── Write results back to workbook ──────────────────────────────────
        for idx, row in enumerate(df.itertuples(index=False), start=2):
            ws.cell(row=idx, column=1).value = row[0]  # RPL-TYPE
            ws.cell(row=idx, column=2).value = row[1]  # RPL-TYPES
            if hyperlinks[idx - 2]:
                ws.cell(row=idx, column=2).hyperlink = hyperlinks[idx - 2]
            ws.cell(row=idx, column=3).value = row[2]  # Previous Data
            ws.cell(row=idx, column=4).value = row[3]  # Current Data
            ws.cell(row=idx, column=5).value = row[4]  # Status
            ws.cell(row=idx, column=6).value = row[5]  # Tracking Number
            ws.cell(row=idx, column=7).value = row[6]  # Creation Date

        # ── Summary Metrics ─────────────────────────────────────────────────
        st.markdown("---")
        st.subheader("📊 Results Summary")

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("🟢 Updated",     len(df[df["Status"] == "Updated"]))
        c2.metric("🔵 Not Updated", len(df[df["Status"] == "Not Updated"]))
        c3.metric("🔴 Failed",      len(df[df["Status"] == "Failed"]))
        c4.metric("📋 Total",       len(df))

        st.dataframe(
            df[["RPL-TYPE", "Previous Data", "Current Data", "Status"]].style.applymap(
                style_status, subset=["Status"]
            ),
            use_container_width=True
        )

        # ── Download Updated Excel ──────────────────────────────────────────
        st.markdown("---")
        st.subheader("📥 Step 3 — Download Updated Excel")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        st.download_button(
            label="⬇️ Download Updated RPS-BOT.xlsx",
            data=output,
            file_name=f"RPS-BOT-{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )

else:
    st.info("👆 Please upload your RPS-BOT.xlsx file to get started.")
    st.markdown("---")
    st.markdown("""
    ### How to use:
    1. 📤 **Upload** your `RPS-BOT.xlsx` file above
    2. 🚀 **Click** the Run Scrapers button
    3. ⏳ **Wait** for all sources to be checked
    4. ⬇️ **Download** the updated Excel file
    """)
